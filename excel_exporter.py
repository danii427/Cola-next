"""
excel_exporter.py - Formatted Multi-Tab Executive Excel Workbook Generator
Generates corporate-styled Excel spreadsheets with summary KPI sheets, scorecards, and clean datasets.
"""

from io import BytesIO
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, Optional


def create_executive_excel_workbook(
    df: pd.DataFrame,
    team_stats: Dict[str, Any],
    thresholds: Dict[str, Any]
) -> bytes:
    """
    Creates a styled multi-tab executive workbook with Summary Dashboard,
    Leaderboard Scorecards, and Full Dataset.
    """
    buffer = BytesIO()
    wb = openpyxl.Workbook()

    # Style Definitions
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=12, bold=True, color="0B1E36")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="1E293B")
    font_regular = Font(name="Calibri", size=10, color="1E293B")

    fill_navy = PatternFill(start_color="0B1E36", end_color="0B1E36", fill_type="solid")
    fill_crimson = PatternFill(start_color="D8232A", end_color="D8232A", fill_type="solid")
    fill_header = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")
    fill_light_gray = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_green = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    fill_yellow = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    fill_red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    # -------------------------------------------------------------
    # SHEET 1: Executive Overview
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_summary.merge_cells("A1:G2")
    cell_title = ws_summary["A1"]
    cell_title.value = "COLA NEXT - SMO KPI EXECUTIVE INTELLIGENCE REPORT"
    cell_title.font = font_title
    cell_title.fill = fill_navy
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    # Team Metric Highlights
    ws_summary["A4"] = "TEAM PERFORMANCE OVERVIEW"
    ws_summary["A4"].font = font_section

    summary_kpis = [
        ("Total Active SMOs", team_stats.get("total_smos", 0), "SMOs"),
        ("Total Sales Target", team_stats.get("total_target", 0), "Cases"),
        ("Total Realized Sales", team_stats.get("total_sales", 0), "Cases"),
        ("Team Overall Achievement", f"{team_stats.get('overall_team_ach', 0):.1f}%", "% Target"),
        ("Avg. Call Completion", f"{team_stats.get('avg_call_comp', 0):.1f}%", "Avg %"),
        ("Avg. Strike Rate", f"{team_stats.get('avg_strike_rate', 0):.1f}%", "Avg %"),
        ("Avg. GPS Accuracy (PJP)", f"{team_stats.get('avg_gps_accuracy', 0):.1f}%", "Avg %"),
        ("Avg. Delivered Cases %", f"{team_stats.get('avg_delivery_pct', 0):.1f}%", "Avg %")
    ]

    ws_summary.append(["KPI Metric", "Realized Value", "Unit / Context"])
    for col_idx in range(1, 4):
        c = ws_summary.cell(row=5, column=col_idx)
        c.font = font_header
        c.fill = fill_header
        c.alignment = Alignment(horizontal="left")

    curr_r = 6
    for kpi, val, unit in summary_kpis:
        ws_summary.append([kpi, val, unit])
        for c_i in range(1, 4):
            cell = ws_summary.cell(row=curr_r, column=c_i)
            cell.font = font_bold if c_i == 2 else font_regular
            cell.border = thin_border
        curr_r += 1

    # Zone Summary
    ws_summary.cell(row=curr_r + 2, column=1, value="HEALTH DISTRIBUTION (ACHIEVEMENT)").font = font_section
    ws_summary.cell(row=curr_r + 3, column=1, value="Green Zone (Met Target)").font = font_bold
    ws_summary.cell(row=curr_r + 3, column=2, value=team_stats.get("green_count", 0)).font = font_regular
    ws_summary.cell(row=curr_r + 4, column=1, value="Yellow Zone (Acceptable)").font = font_bold
    ws_summary.cell(row=curr_r + 4, column=2, value=team_stats.get("yellow_count", 0)).font = font_regular
    ws_summary.cell(row=curr_r + 5, column=1, value="Red Zone (Critical)").font = font_bold
    ws_summary.cell(row=curr_r + 5, column=2, value=team_stats.get("red_count", 0)).font = font_regular

    for r_idx in range(curr_r + 3, curr_r + 6):
        for c_idx in range(1, 3):
            ws_summary.cell(row=r_idx, column=c_idx).border = thin_border

    # -------------------------------------------------------------
    # SHEET 2: SMO Leaderboard Scorecards
    # -------------------------------------------------------------
    ws_board = wb.create_sheet(title="SMO Scorecard Leaderboard")
    ws_board.views.sheetView[0].showGridLines = True

    scorecard_cols = [
        "Rank", "SMO Name", "Route Name", "Region", "Zone",
        "Target", "Route Sale", "Ach.%", "Call Comp. %", "Strike Rate %",
        "Delivered Cases %", "GPS Accuracy % (PJP)", "Working Days", "Performance Tier"
    ]

    ws_board.append(scorecard_cols)
    for c_idx in range(1, len(scorecard_cols) + 1):
        cell = ws_board.cell(row=1, column=c_idx)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal="center")

    sorted_df = df.sort_values(by="Ach.%", ascending=False).reset_index(drop=True)
    for i, row in sorted_df.iterrows():
        row_vals = [
            i + 1,
            row.get("SMO Name", ""),
            row.get("Route Name", ""),
            row.get("Region", ""),
            row.get("Zone", ""),
            row.get("Target", 0),
            row.get("Route Sale", 0),
            f"{float(row.get('Ach.%', 0) or 0):.1f}%",
            f"{float(row.get('Call Comp. %', 0) or 0):.1f}%",
            f"{float(row.get('Strike Rate %', 0) or 0):.1f}%",
            f"{float(row.get('Delivered Cases %', 0) or 0):.1f}%",
            f"{float(row.get('GPS Accuracy % (PJP)', 0) or 0):.1f}%",
            row.get("Working Days", 0),
            row.get("Performance Tier", "")
        ]
        ws_board.append(row_vals)
        r_num = i + 2
        for col_idx in range(1, len(scorecard_cols) + 1):
            c = ws_board.cell(row=r_num, column=col_idx)
            c.font = font_regular
            c.border = thin_border
            if r_num % 2 == 1:
                c.fill = fill_light_gray

    # -------------------------------------------------------------
    # SHEET 3: Complete Raw Cleaned Data
    # -------------------------------------------------------------
    ws_data = wb.create_sheet(title="Full KPI Dataset")
    ws_data.views.sheetView[0].showGridLines = True

    all_cols = list(df.columns)
    ws_data.append(all_cols)
    for col_idx in range(1, len(all_cols) + 1):
        cell = ws_data.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header

    for r_i, row in df.iterrows():
        ws_data.append(list(row.values))
        for col_idx in range(1, len(all_cols) + 1):
            ws_data.cell(row=r_i + 2, column=col_idx).font = font_regular

    # Auto-adjust column widths
    for ws in [ws_summary, ws_board, ws_data]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
